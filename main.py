import streamlit as st
import pandas as pd
import requests
from rapidfuzz import fuzz
import unidecode
from io import BytesIO
import openpyxl
from decouple import config

# Configuración del token de Canvas
API_TOKEN = config("TOKEN")
CANVAS_API_URL = 'https://canvas.uautonoma.cl/api/v1/'

st.set_page_config(page_title="Generador de reporte de asistencia", page_icon="😔")

headers = {
    'Authorization': f'Bearer {API_TOKEN}'
}

# Función para obtener la lista de estudiantes del curso
def get_students(course_id):
    url = f"{CANVAS_API_URL}courses/{course_id}/users"
    params = {'enrollment_type': 'student', 'per_page': 100}
    students = []
    while url:
        response = requests.get(url, headers=headers, params=params)
        response.raise_for_status()
        students.extend(response.json())
        url = response.links.get('next', {}).get('url')
    return students

# Función para obtener información del curso
def get_course_info(course_id):
    url = f"{CANVAS_API_URL}courses/{course_id}"
    response = requests.get(url, headers=headers)
    response.raise_for_status()
    return response.json()

# Función para buscar la mejor coincidencia entre nombres
def buscar_mejor_coincidencia(csv_name, canvas_names):
    csv_parts = set(unidecode.unidecode(csv_name.lower()).split())
    best_match = None
    highest_similarity = 0

    for canvas_name in canvas_names:
        canvas_parts = set(unidecode.unidecode(canvas_name.lower()).split())

        if len(csv_parts) == 1:
            if any(word in canvas_parts for word in csv_parts):
                similarity = fuzz.token_sort_ratio(csv_name, canvas_name)
                if similarity > highest_similarity:
                    highest_similarity = similarity
                    best_match = canvas_name
        else:
            if csv_parts.issubset(canvas_parts):
                similarity = fuzz.token_sort_ratio(csv_name, canvas_name)
                if similarity > highest_similarity:
                    highest_similarity = similarity
                    best_match = canvas_name

    return best_match, highest_similarity

# Interfaz de usuario con Streamlit
st.title("Generador de reporte de asistencia. (Version BETA) 😔")

# Subir archivo CSV
uploaded_file = st.file_uploader("Sube el archivo CSV que te entrega el reporte de Zoom", type=["csv"])

# Ingresar el ID del curso
course_id = st.text_input("Ingresa el ID del curso con el que quieres comparar la asistencia")

if uploaded_file and course_id:
    # Leer el archivo CSV y limpiar duplicados
    info = pd.read_csv(uploaded_file)
    sin_duplicados = info.drop_duplicates(subset='Nombre de usuario', keep='first')
    names_csv = sin_duplicados['Nombre de usuario'].tolist()

    # Obtener nombres de estudiantes desde Canvas
    with st.spinner("Obteniendo estudiantes desde Canvas..."):
        students = get_students(course_id)

    # Obtener información del curso
    with st.spinner("Obteniendo información del curso..."):
        course_info = get_course_info(course_id)
        course_name = course_info['name']
        subaccount_name = course_info['account_id']

    # Comparar nombres del CSV con estudiantes de Canvas
    used_names = set()
    matches = []
    for student in students:
        sortable_name = student['sortable_name']
        sortable_parts = sortable_name.split(",")
        sortable_first = sortable_parts[0].strip() if len(sortable_parts) > 0 else ""
        sortable_last = sortable_parts[1].strip() if len(sortable_parts) > 1 else ""

        matched_name = None
        participation = "No Participó"  # Inicializar con "No Participó"
        for csv_name in names_csv:
            best_match, similarity = buscar_mejor_coincidencia(csv_name, [student['name']])
            if similarity >= 50:
                matched_name = csv_name
                participation = "✔️"  # Si hubo coincidencia, marcar como "Participó"
                used_names.add(csv_name)
                break
        matches.append((sortable_last, sortable_first, matched_name, participation))

    # Crear el DataFrame con los resultados
    results_df = pd.DataFrame(matches, columns=['Nombres', 'Apellido', 'Nombre en CSV', 'Participacion'])

    # Asegurarse de que todas las filas de "Participación" tienen un valor
    results_df['Participacion'] = results_df['Participacion'].apply(lambda x: x if x in ['✔️', '❌'] else '❌')
    
    # Mostrar la tabla interactiva con checkboxes
    st.subheader("Resultados")
    edited_df = st.data_editor(results_df, column_config={
        'Participacion': st.column_config.SelectboxColumn(
            "Participación",
            help="Marca si el estudiante participó",
            options=["✔️", "❌"],
            width="large",
            required=True
        )
    })

    # Nombres del CSV que no se usaron
    unused_names = [name for name in names_csv if name not in used_names]

    # Mostrar nombres del CSV no utilizados
    st.subheader("Nombres sin match en el CSV")
    st.write(unused_names)

    # Configurar el archivo Excel con formato requerido
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        workbook = writer.book
        sheet_name = "Asistencia"
        worksheet = workbook.create_sheet(sheet_name)
        worksheet.append([f"Diplomado: {course_name} (ID: {course_id})"])
        worksheet.append([f"Subcuenta: {subaccount_name}"])
        worksheet.append([])  # Fila en blanco

        # Escribir los datos del DataFrame
        for row in edited_df.itertuples(index=False, name=None):
            worksheet.append(row)

        # Aplicar formato en negrita para las primeras dos filas
        bold_font = openpyxl.styles.Font(bold=True)
        worksheet["A1"].font = bold_font
        worksheet["A2"].font = bold_font

        # Ajustar ancho de columnas
        for col in worksheet.columns:
            col_letter = col[0].column_letter
            worksheet.column_dimensions[col_letter].width = 30

        # Colorear las celdas según participación y añadir bordes
        thin_border = openpyxl.styles.Border(
            left=openpyxl.styles.Side(style='thin'),
            right=openpyxl.styles.Side(style='thin'),
            top=openpyxl.styles.Side(style='thin'),
            bottom=openpyxl.styles.Side(style='thin')
        )

        for row in worksheet.iter_rows(min_row=4, max_row=worksheet.max_row, min_col=1, max_col=4):
            for cell in row:
                cell.border = thin_border
                if cell.column_letter == 'D':
                    if cell.value == "✔️":
                        cell.value = "Participo"
                        cell.fill = openpyxl.styles.PatternFill(start_color="27ae60", end_color="27ae60", fill_type="solid")
                    elif cell.value == "❌":
                        cell.value = "No Participo"
                        cell.fill = openpyxl.styles.PatternFill(start_color="cb4821", end_color="cb4821", fill_type="solid")

    output.seek(0)

    # Botón para descargar el archivo Excel
    st.download_button(
        label="Descargar Reporte",
        data=output,
        file_name=f"{course_name}-{course_id}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
